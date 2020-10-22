from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
def move_area_to_another_excel(start_path,end_path,area_hang,area_end_lie,xieruhang):
    wb_end_path=load_workbook(end_path)
    list_area=[]
    ws_start_path=load_workbook(start_path).active
    for row in ws_start_path['B5':str(area_end_lie)+'{0}'.format(4+area_hang)]:
        list_area.append(row)
    area_end_lie=column_index_from_string(area_end_lie)
    ws_end_path=wb_end_path.active
    for row in range(xieruhang,area_hang+xieruhang):
        for col in range(2,area_end_lie+1):
            if row == xieruhang:
                ws_end_path.cell(row=xieruhang, column=col).value = list_area[0][col - 2].value
            if row == xieruhang+1:
                ws_end_path.cell(row=xieruhang+1, column=col).value = list_area[1][col - 2].value
            if row == xieruhang+2:
                ws_end_path.cell(row=xieruhang+2, column=col).value = list_area[2][col - 2].value
            if row == xieruhang+3:
                ws_end_path.cell(row=xieruhang+3, column=col).value = list_area[3][col - 2].value
            if row == xieruhang+4:
                ws_end_path.cell(row=xieruhang+4, column=col).value = list_area[4][col - 2].value
            if row == xieruhang+5:
                ws_end_path.cell(row=xieruhang+5, column=col).value = list_area[5][col - 2].value
            # if row == 12:
            #     ws_end_path.cell(row=12, column=col).value = list_area[6][col - 2].value
    wb_end_path.save('汇总'+start_path)

shizhongqu_xiangmushu=int(input('请输入市中区有多少个项目：'))
shizhongqu_lieshu=input('请输入市中区表格中最后一列有效数据列数。如AJ：')
move_area_to_another_excel('市中区.xlsx','总表.xlsx',shizhongqu_xiangmushu,shizhongqu_lieshu,6)

dongxinqu_xiangmushu=int(input('请输入东兴区有多少个项目：'))
dongxinqu_lieshu=input('请输入东兴区表格中最后一列有效数据列数。如AJ：')
move_area_to_another_excel('东兴区.xlsx','汇总市中区.xlsx',dongxinqu_xiangmushu,dongxinqu_lieshu,6+shizhongqu_xiangmushu)

longchangshi_xiangmushu=int(input('请输入隆昌市有多少个项目：'))
longchangshi_lieshu=input('请输入隆昌市表格中最后一列有效数据列数。如AJ：')
move_area_to_another_excel('隆昌市.xlsx','汇总东兴区.xlsx',longchangshi_xiangmushu,longchangshi_lieshu,6+shizhongqu_xiangmushu+dongxinqu_xiangmushu)

zizhong_xiangmushu=int(input('请输入资中有多少个项目：'))
zizhong_lieshu=input('请输入资中表格中最后一列有效数据列数。如AJ：')
move_area_to_another_excel('资中.xlsx','汇总隆昌市.xlsx',zizhong_xiangmushu,zizhong_lieshu,6+shizhongqu_xiangmushu+dongxinqu_xiangmushu+longchangshi_xiangmushu)




# print(len(list_shizhongqu))
# print(list_shizhongqu)
# print(list_shizhongqu[0])
# print(list_shizhongqu[0][0])
# print(list_shizhongqu[0][0].value)