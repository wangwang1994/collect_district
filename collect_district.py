from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

list_shizhongqu=[]
wb_shizhongqu=load_workbook('市中区.xlsx')
shizhongqu_xiangmushu=int(input('请输入市中区有多少个项目：'))
shizhongqu_lieshu=input('请输入市中区表格中最后一列有效数据列数。如AJ：')
ws_shizhongqu=wb_shizhongqu.active
for row in ws_shizhongqu['B5':str(shizhongqu_lieshu)+'{0}'.format(4+shizhongqu_xiangmushu)]:
    list_shizhongqu.append(row)
shizhongqu_lieshu=column_index_from_string(shizhongqu_lieshu)
wb_zongbiao=load_workbook('总表.xlsx')
ws_zongbiao=wb_zongbiao.active
for row in range(6,shizhongqu_xiangmushu+6):
    for col in range(2,shizhongqu_lieshu+1):
        if row==6:
            ws_zongbiao.cell(row=6,column=col).value=list_shizhongqu[0][col-2].value
        if row==7:
            ws_zongbiao.cell(row=7, column=col).value = list_shizhongqu[1][col-2].value
        if row == 8:
            ws_zongbiao.cell(row=8, column=col).value = list_shizhongqu[2][col-2].value
# wb_zongbiao.save('总表_市中区添加后.xlsx')
# 下面是小计的功能实现
# ws_zongbiao['B{0}'.format(6+shizhongqu_xiangmushu)]='小计'
# shizhongqu_xiangmushu_zonghe=0
# for i in range(shizhongqu_xiangmushu):
#     shizhongqu_xiangmushu_zonghe=shizhongqu_xiangmushu_zonghe+list_shizhongqu[2+21*i]
# ws_zongbiao['D{0}'.format(6+shizhongqu_xiangmushu)]=shizhongqu_xiangmushu_zonghe
list_dongxinqu=[]
wb_dongxinqu=load_workbook('东兴区.xlsx')
dongxinqu_xiangmushu=int(input('请输入东兴区有多少个项目：'))
dongxinqu_lieshu=input('请输入东兴区表格中最后一列有效数据列数。如AJ：')
ws_dongxinqu=wb_dongxinqu.active
for row in ws_dongxinqu['B5':str(dongxinqu_lieshu)+'{0}'.format(4+dongxinqu_xiangmushu)]:
    list_dongxinqu.append(row)
# print(len(list_dongxinqu))
# print(list_dongxinqu)
for row in range(shizhongqu_xiangmushu+6,dongxinqu_xiangmushu+shizhongqu_xiangmushu+6):
    for col in range(2,dongxinqu_lieshu+1):
        if row==shizhongqu_xiangmushu+6:
            ws_zongbiao.cell(row=shizhongqu_xiangmushu+6,column=col).value=list_shizhongqu[0][col-2].value
        if row==shizhongqu_xiangmushu+7:
            ws_zongbiao.cell(row=shizhongqu_xiangmushu+7, column=col).value =list_shizhongqu[1][col-2].value
        if row == shizhongqu_xiangmushu+8:
            ws_zongbiao.cell(row=shizhongqu_xiangmushu+8, column=col).value = list_shizhongqu[2][col-2].value
        if row == shizhongqu_xiangmushu+9:
            ws_zongbiao.cell(row=shizhongqu_xiangmushu+9, column=col).value = list_shizhongqu[3][col-2].value
        if row == shizhongqu_xiangmushu+10:
            ws_zongbiao.cell(row=shizhongqu_xiangmushu+10, column=col).value = list_shizhongqu[4][col-2].value
        if row == shizhongqu_xiangmushu+11:
            ws_zongbiao.cell(row=shizhongqu_xiangmushu+11, column=col).value = list_dongxinqu[col - 2+21*5]

list_longchangshi=[]
wb_longchangshi=load_workbook('隆昌市.xlsx')
longchangshi_xiangmushu=int(input('请输入隆昌市有多少个项目：'))
ws_longchangshi=wb_longchangshi.active
for row in ws_longchangshi['B5':'V{0}'.format(4+longchangshi_xiangmushu)]:
    for cell in row:
        list_longchangshi.append(cell.value)
# print(len(list_longchangshi))
# print(list_longchangshi)
for row in range(dongxinqu_xiangmushu+shizhongqu_xiangmushu+6,longchangshi_xiangmushu+dongxinqu_xiangmushu+shizhongqu_xiangmushu+6):
    for col in range(2,23):
        if row==dongxinqu_xiangmushu+shizhongqu_xiangmushu+6:
            ws_zongbiao.cell(row=dongxinqu_xiangmushu+shizhongqu_xiangmushu+6,column=col).value=list_longchangshi[col-2]
        if row==dongxinqu_xiangmushu+shizhongqu_xiangmushu+7:
            ws_zongbiao.cell(row=dongxinqu_xiangmushu+shizhongqu_xiangmushu+7, column=col).value = list_longchangshi[col - 2+21]
        if row == dongxinqu_xiangmushu+shizhongqu_xiangmushu+8:
            ws_zongbiao.cell(row=dongxinqu_xiangmushu+shizhongqu_xiangmushu+8, column=col).value = list_longchangshi[col - 2+21*2]
        if row == dongxinqu_xiangmushu+shizhongqu_xiangmushu+9:
            ws_zongbiao.cell(row=dongxinqu_xiangmushu+shizhongqu_xiangmushu+9, column=col).value = list_longchangshi[col - 2+21*3]
        if row == dongxinqu_xiangmushu+shizhongqu_xiangmushu+10:
            ws_zongbiao.cell(row=dongxinqu_xiangmushu+shizhongqu_xiangmushu+10, column=col).value = list_longchangshi[col - 2+21*4]
        if row == dongxinqu_xiangmushu+shizhongqu_xiangmushu+11:
            ws_zongbiao.cell(row=dongxinqu_xiangmushu+shizhongqu_xiangmushu+11, column=col).value = list_longchangshi[col - 2+21*5]

wb_zongbiao.save('总表_添加后.xlsx')



# print(list_shizhongqu)

